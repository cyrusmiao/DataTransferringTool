import pandas as pd
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from thefuzz import fuzz
from typing import List, Dict, Any
import xlwt
from .config import TransferConfig

class DataTransfer:
    def __init__(self, config: TransferConfig):
        self.config = config
        self.report = []
        self.target_workbook = None
        self.target_sheet_name = None
        self.conflict_cell_counts = {}
        self.target_df = self._load_target_file()
        
    def _col_to_index(self, col) -> int:
        if isinstance(col, int):
            return col
        col = str(col).strip()
        if col.isdigit():
            return int(col)
        num = 0
        for c in col.upper():
            if c.isalpha():
                num = num * 26 + (ord(c) - ord('A')) + 1
        return num - 1

    def _index_to_col(self, index: int) -> str:
        if index < 0:
            raise ValueError("Column index must be non-negative.")
        label = ""
        index += 1
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            label = chr(ord('A') + remainder) + label
        return label
        
    def _resolve_sheet_name(self, workbook: dict[str, pd.DataFrame], sheet_name: str | int | None) -> str:
        sheet_names = list(workbook.keys())
        if not sheet_names:
            raise ValueError("Excel file does not contain any sheets.")
        if sheet_name is None:
            return sheet_names[0]
        if isinstance(sheet_name, int):
            if sheet_name < 0 or sheet_name >= len(sheet_names):
                raise ValueError(f"Sheet index '{sheet_name}' is out of bounds.")
            return sheet_names[sheet_name]
        normalized_sheet_name = str(sheet_name).strip()
        for actual_sheet_name in sheet_names:
            if actual_sheet_name == sheet_name:
                return actual_sheet_name
            if str(actual_sheet_name).strip() == normalized_sheet_name:
                return actual_sheet_name
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")

    def _load_workbook(self, file_path: str) -> dict[str, pd.DataFrame]:
        return pd.read_excel(Path(file_path), sheet_name=None)

    def _load_sheet_file(self, file_path: str, sheet_name: str | int | None = None) -> tuple[pd.DataFrame, str | None]:
        path = Path(file_path)
        if path.suffix.lower() == '.csv':
            return pd.read_csv(path), None
        if path.suffix.lower() in ['.xls', '.xlsx']:
            workbook = self._load_workbook(file_path)
            resolved_sheet_name = self._resolve_sheet_name(workbook, sheet_name)
            return workbook[resolved_sheet_name].copy(), resolved_sheet_name
        raise ValueError(f"Unsupported file format: {path.suffix}")

    def _load_target_file(self) -> pd.DataFrame:
        path = Path(self.config.target_file)
        if path.suffix.lower() in ['.xls', '.xlsx']:
            self.target_workbook = self._load_workbook(self.config.target_file)
            self.target_sheet_name = self._resolve_sheet_name(self.target_workbook, self.config.target_sheet)
            return self.target_workbook[self.target_sheet_name].copy()
        return self._load_file(self.config.target_file)

    def _load_file(self, file_path: str, sheet_name: str | int | None = None) -> pd.DataFrame:
        path = Path(file_path)
        if path.suffix.lower() == '.csv':
            return pd.read_csv(path)
        elif path.suffix.lower() in ['.xls', '.xlsx']:
            if sheet_name is None:
                return pd.read_excel(path)
            return pd.read_excel(path, sheet_name=sheet_name)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")

    def _prepare_value_for_target(self, df: pd.DataFrame, column: str, value):
        if pd.isna(value):
            return value
        if pd.api.types.is_string_dtype(df[column].dtype):
            if isinstance(value, str):
                return value
            try:
                df[column] = df[column].astype('object')
                return pd.to_numeric([value], errors='raise')[0]
            except (TypeError, ValueError):
                return str(value)
        if pd.api.types.is_numeric_dtype(df[column].dtype):
            try:
                return pd.to_numeric([value], errors='raise')[0]
            except (TypeError, ValueError):
                df[column] = df[column].astype('object')
        return value

    def _normalize_comparison_value(self, value):
        if pd.isna(value):
            return None
        if isinstance(value, str):
            value = value.strip()
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return value

    def _values_are_equivalent(self, old_val, new_val) -> bool:
        return self._normalize_comparison_value(old_val) == self._normalize_comparison_value(new_val)

    def _display_header(self, header):
        header_text = str(header)
        if header_text.startswith("Unnamed:"):
            return ""
        return header

    def _prepare_df_for_export(self, df: pd.DataFrame) -> pd.DataFrame:
        export_df = df.copy()
        export_df.columns = [self._display_header(column) for column in export_df.columns]
        return export_df

    def _coerce_excel_scalar(self, value):
        if pd.isna(value):
            return None
        if hasattr(value, "item") and not isinstance(value, (str, bytes)):
            try:
                value = value.item()
            except (ValueError, TypeError):
                pass
        if isinstance(value, Decimal):
            if value == value.to_integral_value():
                return int(value)
            return float(value)
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value

    def _highlight_conflict_cells_enabled(self, output_path: Path) -> bool:
        return self.config.highlight_conflict_cells and output_path.suffix.lower() in ['.xls', '.xlsx']

    def _build_xls_cell_style(self, conflict_count: int | None = None, is_datetime: bool = False, is_date: bool = False):
        style = xlwt.XFStyle()
        if is_datetime:
            style.num_format_str = 'YYYY-MM-DD HH:MM:SS'
        elif is_date:
            style.num_format_str = 'YYYY-MM-DD'
        if conflict_count:
            pattern = xlwt.Pattern()
            pattern.pattern = xlwt.Pattern.SOLID_PATTERN
            if conflict_count == 1:
                pattern.pattern_fore_colour = 5
            elif conflict_count == 2:
                pattern.pattern_fore_colour = 0x21
            else:
                pattern.pattern_fore_colour = 10
            style.pattern = pattern
        return style

    def _get_conflict_fill(self, conflict_count: int) -> PatternFill:
        if conflict_count == 1:
            color = "FFFF00"
        elif conflict_count == 2:
            color = "FFC000"
        else:
            color = "FF0000"
        return PatternFill(fill_type="solid", start_color=color, end_color=color)

    def _record_conflict_cell(self, row_index: int, column_index: int):
        key = (row_index, column_index)
        self.conflict_cell_counts[key] = self.conflict_cell_counts.get(key, 0) + 1

    def _apply_xlsx_conflict_highlights(self, path: Path, df: pd.DataFrame):
        if not self.conflict_cell_counts:
            return
        workbook = load_workbook(path)
        worksheet = workbook[self.target_sheet_name if self.target_sheet_name is not None else workbook.sheetnames[0]]
        for (row_index, column_index), conflict_count in self.conflict_cell_counts.items():
            excel_row = row_index + 2
            excel_col = column_index + 1
            worksheet.cell(row=excel_row, column=excel_col).fill = self._get_conflict_fill(conflict_count)
        workbook.save(path)

    def _save_xls_workbook(self, workbook_data: dict[str, pd.DataFrame], path: Path):
        workbook = xlwt.Workbook(encoding='utf-8')
        workbook.set_colour_RGB(0x21, 255, 192, 0)
        for sheet_name, sheet_df in workbook_data.items():
            worksheet = workbook.add_sheet(str(sheet_name)[:31])
            for col_idx, column in enumerate(sheet_df.columns):
                worksheet.write(0, col_idx, self._coerce_excel_scalar(column))
            for row_idx, row in enumerate(sheet_df.itertuples(index=False, name=None), start=1):
                for col_idx, value in enumerate(row):
                    excel_value = self._coerce_excel_scalar(value)
                    if excel_value is None:
                        continue
                    conflict_count = None
                    if self._highlight_conflict_cells_enabled(path) and sheet_name == self.target_sheet_name:
                        conflict_count = self.conflict_cell_counts.get((row_idx - 1, col_idx))
                    style = self._build_xls_cell_style(
                        conflict_count=conflict_count,
                        is_datetime=isinstance(excel_value, datetime),
                        is_date=isinstance(excel_value, date) and not isinstance(excel_value, datetime)
                    )
                    worksheet.write(row_idx, col_idx, excel_value, style)
        workbook.save(str(path))

    def _save_file(self, df: pd.DataFrame, file_path: str):
        path = Path(file_path)
        if path.suffix.lower() == '.csv':
            self._prepare_df_for_export(df).to_csv(path, index=False)
        elif path.suffix.lower() == '.xls':
            if self.target_workbook is not None:
                workbook_data = {
                    sheet: self._prepare_df_for_export(sheet_df)
                    for sheet, sheet_df in self.target_workbook.items()
                }
                workbook_data[self.target_sheet_name] = self._prepare_df_for_export(df)
            else:
                workbook_data = {"Sheet1": self._prepare_df_for_export(df)}
            self._save_xls_workbook(workbook_data, path)
        elif path.suffix.lower() == '.xlsx':
            if self.target_workbook is not None:
                workbook = {
                    sheet: self._prepare_df_for_export(sheet_df)
                    for sheet, sheet_df in self.target_workbook.items()
                }
                workbook[self.target_sheet_name] = self._prepare_df_for_export(df)
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    for sheet, sheet_df in workbook.items():
                        sheet_df.to_excel(writer, index=False, sheet_name=sheet)
            else:
                self._prepare_df_for_export(df).to_excel(path, index=False, engine='openpyxl')
            if self._highlight_conflict_cells_enabled(path):
                self._apply_xlsx_conflict_highlights(path, df)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")

    def _interactive_resolve(self, old_val, new_val, src_file, target_file, ref_val) -> str:
        print(f"\nConflict Detected!")
        print(f"Target file: {target_file}")
        print(f"Source file: {src_file}")
        print(f"Reference value: {ref_val}")
        print(f"Original Data: {old_val}")
        print(f"New Data: {new_val}")
        print("Options: [1] Keep Original [2] Overwrite")
        while True:
            choice = input("Enter choice (1/2): ").strip()
            if choice == '1':
                return 'keep_original'
            elif choice == '2':
                return 'overwrite'
            else:
                print("Invalid choice, try again.")

    def run(self):
        # We need to process each source
        # Make a copy of the target to modify
        out_df = self.target_df.copy()
        
        for source_idx, source in enumerate(self.config.sources):
            try:
                src_df, source_sheet_name = self._load_sheet_file(source.file_path, source.sheet_name)
            except Exception as e:
                print(f"Error loading source file {source.file_path}: {e}")
                continue
                
            if not source.reference_column:
                print(f"Warning: No reference column defined for {source.file_path}. Skipping.")
                continue

            ref_src_key, ref_tgt_key = list(source.reference_column.items())[0]
            ref_src_idx = self._col_to_index(ref_src_key)
            ref_tgt_idx = self._col_to_index(ref_tgt_key)
            
            if ref_src_idx >= len(src_df.columns):
                print(f"Warning: Reference column '{ref_src_key}' out of bounds in {source.file_path}. Skipping source.")
                continue
            if ref_tgt_idx >= len(out_df.columns):
                print(f"Warning: Reference column '{ref_tgt_key}' out of bounds in target file. Skipping source.")
                continue
                
            ref_col_src = src_df.columns[ref_src_idx]
            ref_col_tgt = out_df.columns[ref_tgt_idx]
            
            # Map target dataframe to dictionary of indices for quick lookup by reference column
            # Since target might have duplicate reference values, we store a list of indices
            tgt_dict = {}
            for target_idx, ref_val in out_df[ref_col_tgt].items():
                if pd.isna(ref_val):
                    continue
                if ref_val not in tgt_dict:
                    tgt_dict[ref_val] = []
                tgt_dict[ref_val].append(target_idx)
                
            # Parse mappings
            valid_mappings = []
            for src_key, tgt_key in source.mapping:
                src_idx = self._col_to_index(src_key)
                tgt_idx = self._col_to_index(tgt_key)
                if src_idx < len(src_df.columns) and tgt_idx < len(out_df.columns):
                    valid_mappings.append({
                        'src_letter': self._index_to_col(src_idx),
                        'src_col': src_df.columns[src_idx],
                        'tgt_idx': tgt_idx,
                        'tgt_col': out_df.columns[tgt_idx],
                        'tgt_letter': self._index_to_col(tgt_idx)
                    })

            for src_idx, src_row in src_df.iterrows():
                ref_val = src_row[ref_col_src]
                if pd.isna(ref_val):
                    continue
                
                if ref_val not in tgt_dict:
                    # Target row not found, report and skip
                    print(f"Report: Source row with reference '{ref_src_key}={ref_val}' from '{source.file_path}' not found in target file '{self.config.target_file}'. Skipping.")
                    self.report.append({
                        "conflict_resolution": "skipped_not_in_target",
                        "source_file": source.file_path,
                        "source_sheet": source_sheet_name,
                        "source_column": self._index_to_col(ref_src_idx),
                        "source_headers": self._display_header(ref_col_src),
                        "target_file": self.config.target_file,
                        "target_sheet": self.target_sheet_name,
                        "target_column": self._index_to_col(ref_tgt_idx),
                        "target_headers": self._display_header(ref_col_tgt),
                        "reference_value": ref_val,
                        "original_data": None,
                        "new_data": None,
                        "similarity_score": None
                    })
                    continue
                
                # For each matching target row, transfer data
                target_indices = tgt_dict[ref_val]
                for target_idx in target_indices:
                    for mapping in valid_mappings:
                        src_col = mapping['src_col']
                        tgt_col = mapping['tgt_col']
                        
                        new_val = src_row[src_col]
                        old_val = out_df.at[target_idx, tgt_col]
                        
                        if pd.isna(new_val):
                            continue # Nothing to transfer

                        prepared_new_val = self._prepare_value_for_target(out_df, tgt_col, new_val)
                        
                        action_taken = "transferred"
                        similarity = 0.0
                        
                        if pd.isna(old_val):
                            out_df.at[target_idx, tgt_col] = prepared_new_val
                        else:
                            # Both have values, calculate similarity
                            similarity = fuzz.ratio(str(old_val), str(prepared_new_val))
                            
                            if self._values_are_equivalent(old_val, prepared_new_val):
                                action_taken = "identical_skipped"
                            else:
                                self._record_conflict_cell(target_idx, mapping['tgt_idx'])
                                resolution = self.config.conflict_resolution
                                if resolution == 'manual':
                                    resolution = self._interactive_resolve(old_val, new_val, source.file_path, self.config.target_file, ref_val)
                                
                                if resolution == 'keep_original':
                                    action_taken = "conflict_kept_original"
                                elif resolution == 'overwrite':
                                    out_df.at[target_idx, tgt_col] = prepared_new_val
                                    action_taken = "conflict_overwritten"
                        
                        self.report.append({
                            "conflict_resolution": action_taken,
                            "source_file": source.file_path,
                            "source_sheet": source_sheet_name,
                            "source_column": mapping['src_letter'],
                            "source_headers": self._display_header(src_col),
                            "target_file": self.config.target_file,
                            "target_sheet": self.target_sheet_name,
                            "target_column": mapping['tgt_letter'],
                            "target_headers": self._display_header(tgt_col),
                            "reference_value": ref_val,
                            "original_data": old_val if not pd.isna(old_val) else None,
                            "new_data": prepared_new_val,
                            "similarity_score": similarity
                        })

        self._save_file(out_df, self.config.output_file)
        if self.config.generate_transfer_report:
            self._generate_report()
        if self.config.generate_reference_report:
            self._generate_reference_report()

    def _generate_report(self):
        report_df = pd.DataFrame(self.report)
        report_path = "transfer_report.xlsx"
        if not report_df.empty:
            report_df.to_excel(report_path, index=False)
            print(f"Report generated successfully: {report_path}")
        else:
            print("No transfers or conflicts to report.")

    def _generate_reference_report(self):
        reference_report_path = Path("reference_report.md")
        grouped_references = {}
        for row in self.report:
            if row["conflict_resolution"] == "skipped_not_in_target":
                continue
            source_file = row["source_file"]
            if source_file not in grouped_references:
                grouped_references[source_file] = []
            if row["reference_value"] not in grouped_references[source_file]:
                grouped_references[source_file].append(row["reference_value"])

        if not grouped_references:
            print("No non-skipped reference values to report.")
            return

        lines = ["# Reference Report", ""]
        for source_file, reference_values in grouped_references.items():
            lines.append(f"## {source_file}")
            lines.append("")
            for reference_value in reference_values:
                lines.append(f"- {reference_value}")
            lines.append("")
        reference_report_path.write_text("\n".join(lines), encoding="utf-8")
        print(f"Reference report generated successfully: {reference_report_path}")
